# -*- coding: utf-8 -*-
# @Time    : 2021/4/22 16:33
# @Author  : derrick
# @Email   : luotaibin@gmail.com
# @File    : excel.py
# @Software: PyCharm
from openpyxl import load_workbook
from openpyxl import Workbook


class excel:
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        self.sheets = self.wb.sheetnames

    # 获取表格的总行数和总列数
    def getRowsClosNum(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def getCellValue(self, row, column):
        cellvalue = self.ws.cell(row=row, column=column).value
        return cellvalue

    # 获取某列的所有值
    def getColValues(self, column):
        rows = self.ws.max_row
        columndata = []
        for i in range(1, rows + 1):
            cellvalue = self.ws.cell(row=i, column=column).value
            columndata.append(cellvalue)
        return columndata

    # 获取某行所有值
    def getRowValues(self, row):
        columns = self.ws.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.ws.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    # 设置某个单元格的值
    def setCellValue(self, row, colunm, cellvalue, j):
        self.sheet = self.sheets[j]
        self.ws = self.wb[self.sheet]
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)

    def readExcel(self, j):
        self.sheet = self.sheets[j]
        self.ws = self.wb[self.sheet]

        params = []
        for i in range(2, self.ws.max_row + 1):
            isRun = self.getCellValue(i, 8)
            if isRun == 'N' or isRun == '' or isRun != 'Y' or isRun == 'None':
                continue
            feature = self.getCellValue(i, 1)
            story = self.getCellValue(i, 2)
            param = self.getCellValue(i, 3)
            url = self.getCellValue(i, 4)
            description = self.getCellValue(i, 5)
            uuid = self.getCellValue(i, 6)
            assertion = self.getCellValue(i, 7)
            tem = {"feature": feature, "story": story, "param": param, "url": url, "description": description,
                   "uuid": uuid, "assertion": assertion, "row": i}
            params.append(tem)
        return params
